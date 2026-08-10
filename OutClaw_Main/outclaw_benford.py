#!/usr/bin/env python3
"""
OutClaw/outclaw_benford.py — Benford's Law Financial Fraud Detector.

LWM INTEGRATION (2026-07-26): Extracted from DeepSeek's Legal War Machine
session (DS LWM.txt). Applies Benford's Law to detect anomalies in
numerical data that suggest fraudulent manipulation.

Benford's Law states that in naturally-occurring numerical data, the
leading digit distribution follows a predictable logarithmic pattern.
Deviations from this pattern strongly suggest cooked books.

Usage:
    from OutClaw.outclaw_benford import BenfordAnalyzer
    analyzer = BenfordAnalyzer()
    result = analyzer.analyze([123.45, 6789.01, 45.67, ...])
    if result["fraud_probability"] > 0.5:
        print("WARNING: Anomalous digit distribution detected")
"""

from __future__ import annotations

import math
from collections import Counter
from typing import Any

# Expected Benford distribution (first-digit probabilities)
BENFORD_DISTRIBUTION: dict[int, float] = {
    1: 0.301,
    2: 0.176,
    3: 0.125,
    4: 0.097,
    5: 0.079,
    6: 0.067,
    7: 0.058,
    8: 0.051,
    9: 0.046,
}


# ═══════════════════════════════════════════════════════
#  OUTCLAW_BENFORD
# ═══════════════════════════════════════════════════════════════

class BenfordAnalyzer:
    """
    Applies Benford's Law to detect financial anomalies.

    Works on any iterable of numbers: transaction amounts, invoice totals,
    campaign contributions, expense reports, contract values.
    """

    def __init__(self, threshold: float = 0.15):
        """
        Args:
            threshold: Maximum allowed deviation from Benford before flagging.
                       Default 0.15 (15% deviation triggers a flag).
        """
        self.threshold = threshold

    def analyze(self, numbers: list[float]) -> dict[str, Any]:
        """
        Analyze a list of numbers for Benford compliance.

        Returns:
            Dict with observed distribution, expected distribution,
            deviations, fraud probability, and flagged digits.
        """
        if len(numbers) < 30:
            return {
                "error": "Insufficient data — Benford's Law requires at least 30 data points for reliability.",
                "sample_size": len(numbers),
            }

        # Extract first digits
        first_digits = [self._first_digit(n) for n in numbers if n > 0]
        if len(first_digits) < 30:
            return {
                "error": "Insufficient non-zero values.",
                "sample_size": len(first_digits),
            }

        # Observed distribution
        observed_counts = Counter(first_digits)
        total = sum(observed_counts.values())
        observed_dist = {d: observed_counts.get(d, 0) / total for d in range(1, 10)}

        # Deviation from expected
        deviations = {}
        max_deviation = 0.0
        flagged_digits = []
        for d in range(1, 10):
            expected = BENFORD_DISTRIBUTION[d]
            actual = observed_dist.get(d, 0)
            deviation = abs(actual - expected)
            deviations[d] = round(deviation, 4)
            if deviation > self.threshold:
                flagged_digits.append(d)
            max_deviation = max(max_deviation, deviation)

        # Fraud probability: weighted by deviation magnitude and count
        fraud_probability = min(
            1.0,
            max_deviation * 3.0 + (len(flagged_digits) / 9) * 0.5,
        )

        # Determine overall verdict
        if fraud_probability > 0.7:
            verdict = "HIGH — strongly suggestive of manipulation"
        elif fraud_probability > 0.4:
            verdict = "MEDIUM — anomalous; warrants investigation"
        elif fraud_probability > 0.15:
            verdict = "LOW — minor deviation; likely benign"
        else:
            verdict = "CLEAN — follows expected Benford distribution"

        return {
            "sample_size": len(first_digits),
            "observed_distribution": {d: round(v, 4) for d, v in observed_dist.items()},
            "expected_distribution": BENFORD_DISTRIBUTION,
            "deviations": deviations,
            "max_deviation": round(max_deviation, 4),
            "flagged_digits": flagged_digits,
            "fraud_probability": round(fraud_probability, 4),
            "verdict": verdict,
        }

    def analyze_from_file(
        self, filepath: str, column: str = "amount"
    ) -> dict[str, Any]:
        """
        Load numbers from a CSV/Excel file and analyze.

        Args:
            filepath: Path to CSV or Excel file.
            column: Column name containing numeric values.
        """
        import csv
        from pathlib import Path

        numbers: list[float] = []
        path = Path(filepath)

        try:
            if path.suffix in (".csv", ".txt"):
                with open(path, "r") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        try:
                            numbers.append(float(row.get(column, 0)))
                        except (ValueError, TypeError):
                            continue
            elif path.suffix in (".xlsx", ".xls"):
                # Requires openpyxl — graceful skip if unavailable
                try:
                    import openpyxl  # type: ignore

                    wb = openpyxl.load_workbook(path)
                    ws = wb.active
                    headers = [
                        cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
                    ]
                    if column in headers:
                        col_idx = headers.index(column)
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            try:
                                val = row[col_idx]
                                if val is not None:
                                    numbers.append(float(val))
                            except (ValueError, TypeError, IndexError):
                                continue
                except ImportError:
                    return {
                        "error": "openpyxl not installed. pip install openpyxl for Excel support."
                    }
        except Exception as e:
            return {"error": f"Failed to read file: {e}"}

        return self.analyze(numbers)

    @staticmethod
    def _first_digit(n: float) -> int:
        """Extract the first non-zero digit."""
        n = abs(n)
        if n == 0:
            return 0
        # Get first digit using log10
        try:
            return int(str(int(n))[0])
        except (ValueError, IndexError):
            # Scientific notation fallback
            exponent = int(math.floor(math.log10(n)))
            denom = 10**exponent
            if denom == 0:
                return 1  # underflow guard
            mantissa = n / denom
            return int(mantissa)


# ---------------------------------------------------------------------------
# Convenience
# ---------------------------------------------------------------------------


def check_benford(numbers: list[float]) -> dict[str, Any]:
    """One-liner: check a list of numbers for Benford compliance."""
    return BenfordAnalyzer().analyze(numbers)


# ---------------------------------------------------------------------------
# Smoke test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Generate naturally-distributed numbers (should pass)
    import random

    natural = [random.lognormvariate(5, 1) for _ in range(100)]

    # Generate fraudulent numbers (biased toward 7-9)
    fraudulent = [random.uniform(700, 999) for _ in range(50)] + [
        random.uniform(70, 99) for _ in range(50)
    ]

    analyzer = BenfordAnalyzer()
    print("=== Benford's Law Analysis ===")
    print()

    nat = analyzer.analyze(natural)
    print(
        f"Natural data: fraud_prob={nat['fraud_probability']:.2f}, verdict={nat['verdict']}"
    )

    fraud = analyzer.analyze(fraudulent)
    print(
        f"Fraudulent data: fraud_prob={fraud['fraud_probability']:.2f}, verdict={fraud['verdict']}"
    )
    print(f"  Flagged digits: {fraud['flagged_digits']}")
